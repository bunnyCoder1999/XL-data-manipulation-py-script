import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# import tensorflow as tf
# import numpy as np

wb = xl.load_workbook("transactions.xlsx") # load xlfile into workbook
sheet = wb["Sheet1"] # load sheet into variable (Sheet1 is the name of the sheet and is case sensitive)
cell = sheet["a1"] #or cell = sheet.cell(1-row,1-col)
print(cell.value) # value in the cell A1


cell = sheet.cell(1,4)
cell.value = "Updated price"

x = 2
while sheet.max_row >= x:
    sheet.cell(x,4).value = sheet.cell(x,3).value * 0.9
    x = x+1

wb.save("transactions2.xlsx")
# wb.close()
values = Reference(sheet,min_row = 2,
                    max_row = sheet.max_row, 
                    min_col = 4, 
                    max_col = 4) # Reference is a class that allows you to reference a range of cells


chart = BarChart() # BarChart is a class that allows you to create a bar chart

chart.add_data(values) # add data to the chart
sheet.add_chart(chart, "e2") # add chart to the sheet
wb.save("transactions2.xlsx") # save the file





